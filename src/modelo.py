"""
Modelo de datos y lógica de negocio para el sistema de cataciones.
Separado de la interfaz para poder testear sin necesidad de UI.
"""
from dataclasses import dataclass, asdict
from datetime import datetime
import openpyxl
import os


ARCHIVO_EXCEL = os.path.join("datos", "cataciones.xlsx")

COLUMNAS = [
    "Fecha", "Catador", "Café", "Origen",
    "Acidez", "Cuerpo", "Aroma", "Dulzor", "Balance",
    "Puntaje Total", "Notas"
]


@dataclass
class Catacion:
    """Representa una catación individual con sus puntajes sensoriales."""
    catador: str
    cafe: str
    origen: str
    acidez: float
    cuerpo: float
    aroma: float
    dulzor: float
    balance: float
    notas: str = ""
    fecha: str = ""

    def __post_init__(self):
        if not self.fecha:
            self.fecha = datetime.now().strftime("%Y-%m-%d %H:%M")
        self._validar()

    def _validar(self):
        """Valida que los puntajes estén en rango 1-10."""
        campos = {
            "Acidez": self.acidez, "Cuerpo": self.cuerpo,
            "Aroma": self.aroma, "Dulzor": self.dulzor,
            "Balance": self.balance
        }
        for nombre, valor in campos.items():
            if not (0 <= valor <= 10):
                raise ValueError(f"{nombre} debe estar entre 0 y 10 (recibido: {valor})")
        if not self.catador.strip():
            raise ValueError("El nombre del catador no puede estar vacío")
        if not self.cafe.strip():
            raise ValueError("El nombre del café no puede estar vacío")

    @property
    def puntaje_total(self) -> float:
        """Calcula el puntaje promedio de los 5 criterios sensoriales."""
        return round(
            (self.acidez + self.cuerpo + self.aroma + self.dulzor + self.balance) / 5,
            2
        )

    def a_fila(self) -> list:
        """Convierte la catación en una fila lista para escribir en Excel."""
        return [
            self.fecha, self.catador, self.cafe, self.origen,
            self.acidez, self.cuerpo, self.aroma, self.dulzor, self.balance,
            self.puntaje_total, self.notas
        ]


class RepositorioCataciones:
    """Maneja la persistencia de cataciones en un archivo Excel."""

    def __init__(self, ruta_archivo: str = ARCHIVO_EXCEL):
        self.ruta_archivo = ruta_archivo
        self._inicializar()

    def _inicializar(self):
        carpeta = os.path.dirname(self.ruta_archivo)
        if carpeta and not os.path.exists(carpeta):
            os.makedirs(carpeta, exist_ok=True)
        if not os.path.exists(self.ruta_archivo):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cataciones"
            ws.append(COLUMNAS)
            wb.save(self.ruta_archivo)

    def guardar(self, catacion: Catacion) -> None:
        wb = openpyxl.load_workbook(self.ruta_archivo)
        ws = wb.active
        ws.append(catacion.a_fila())
        wb.save(self.ruta_archivo)

    def listar_todas(self) -> list[dict]:
        wb = openpyxl.load_workbook(self.ruta_archivo)
        ws = wb.active
        resultados = []
        for fila in ws.iter_rows(min_row=2, values_only=True):
            if fila[0] is None:
                continue
            resultados.append(dict(zip(COLUMNAS, fila)))
        return resultados

    def promedio_por_cafe(self) -> dict[str, float]:
        """Agrupa cataciones por nombre de café y calcula el puntaje promedio."""
        datos = self.listar_todas()
        agrupado: dict[str, list[float]] = {}
        for fila in datos:
            cafe = fila["Café"]
            agrupado.setdefault(cafe, []).append(fila["Puntaje Total"])
        return {
            cafe: round(sum(puntajes) / len(puntajes), 2)
            for cafe, puntajes in agrupado.items()
        }
